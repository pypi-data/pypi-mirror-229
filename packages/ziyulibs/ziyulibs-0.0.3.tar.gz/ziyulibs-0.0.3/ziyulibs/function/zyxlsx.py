##################################################
# 功能说明：
# 更新日期：
# 更新次数：
# 完成情况：
# 执行次数：233
# 执行时间：2023-09-08 11:59:39
##################################################
import datetime
import os
import re
import time
import unittest

from openpyxl import Workbook
from openpyxl.utils import get_column_letter  # , column_index_from_string
from ziyulibs.function.zycommon import updatedescribe

updatedescribe(__file__)

# import openpyxl

# from openpyxl.styles import Font, colors, Alignment


def updatexlsx(workbook, sheetname, header, datalist, sheetindex=0, rowformat=None, cellformat=None):
    sheets = workbook.get_sheet_names()
    # 判断要插入的位置
    if (sheetindex > len(sheets)):
        sheetindex = len(sheets)
    # 如果没有sheet需要新建
    if (sheetname not in sheets):
        workbook.create_sheet(title=sheetname, index=sheetindex)
    worksheet = workbook.get_sheet_by_name(sheetname)
    # 生成表头
    for idx1, row in enumerate(header):
        title = ''
        width = 10

        try:
            if (isinstance(row, str)):
                # 字符串模式切割为:列头名,字段名,宽度
                keys = row.split(':')
                title = keys[0] if len(keys) > 0 else ''
                width = int(keys[1]) if len(keys) > 1 else 10
            if (isinstance(row, dict)):
                title = row['name'] if 'name' in row else ''
                width = int(row['width']) if 'width' in row else 10
        except Exception as ex:
            print(str(ex))

        worksheet.cell(row=1, column=idx1+1).value = title
        if (width != 0):
            worksheet.column_dimensions[get_column_letter(idx1+1)].width = width
    # 生成数据
    for idx1, row in enumerate(datalist):
        rowidx = idx1 + 2
        dataitem = datalist[idx1]
        for idx2, item in enumerate(header):
            colidx = idx2+1
            datacell = worksheet.cell(row=rowidx, column=colidx)
            key = ''
            # 字符串模式切割为:列头名,字段名,宽度
            if (isinstance(item, str)):
                keys = item.split(':')
                if (len(keys) < 2):
                    print('行%(idx1)s,列%(idx2)s缺少key1' % {'idx1': rowidx, 'idx2': colidx})
                    continue
                key = keys[1]
            if (isinstance(item, dict)):
                if ('formula' in item):
                    v = re.sub('#', str(rowidx), item['formula'])
                    datacell.value = v
                    continue
                if ('key' not in item):
                    print('行%(idx1)s,列%(idx2)s缺少key2' % {'idx1': rowidx, 'idx2': colidx})
                    continue
                key = item['key']
            if (key == 'index'):
                datacell.value = idx1+1
                continue
            if (key not in dataitem):
                print('行%(idx1)s,列%(idx2)s缺少value' % {'idx1': rowidx, 'idx2': colidx})
                continue
            value = dataitem[key]
            datacell.value = dataitem[key]
            if (cellformat != None):
                cellformat(datacell, value, rowidx, colidx)
    return workbook


def readxls(workbook, workname, tbconfig, cellformat=None):
    # print(workname)

    sheetnames = workbook.sheet_names()
    print(sheetnames)
    if (workname not in sheetnames):
        print(sheetnames, '不包含', workname)
        return
    worksheet = workbook.sheet_by_name(workname)
    # work_sheet = work_book.sheet_by_index(0)  # 通过表名的索引打开工作表
    rows = worksheet.max_row
    cols = worksheet.max_column
    # print(rows, cols)
    if (tbconfig is None or len(tbconfig) == 0):
        tbconfig = []
        for item in range(cols):
            cell = worksheet.cell(1, item+1).value
            tbconfig.append(cell)
    # 处理 tbconfig 格式
    for idx, item in enumerate(tbconfig):
        if (isinstance(item, str)):
            tbconfig[idx] = {'key': item}

    datalist = []
    for idx1 in range(rows-1):
        dataitem = {}
        for idx2, item2 in enumerate(tbconfig):
            datakey = item2['key']
            datavalue = worksheet.cell(idx1+2, idx2+1).value
            dataitem[datakey] = datavalue
            if (cellformat != None):
                cellformat(worksheet, idx1+2, idx2+1, item2, datavalue, dataitem, datakey)

            if ('colsformat' in item2):
                item2['colsformat'](worksheet, idx1+2, idx2+1, item2, datavalue, dataitem, datakey)
        datalist.append(dataitem)
    return datalist


def writexls(path, datalist, columnmaps):
    try:
        data_book = Workbook()
        data_sheet = data_book.active
        # colName = ['AA', 'BB', 'CC', 'DD', 'EE', 'FF']
        # for col in range(len(colName)):
        #     data_sheet.cell(column=col + 1, row=1, value=colName[col])

        # columnmaps={'idx':'序号','tp':'图片','bt':'标题','lx':'类型'}

        keys = list(columnmaps.keys())
        values = list(columnmaps.values())
        titles = []
        # 处理一下title
        for idx1, value in enumerate(values):
            if (isinstance(value, dict)):
                titles.append(value['name'])
                # print(get_column_letter(idx1+1))
                # print(value['width'])
                data_sheet.column_dimensions[get_column_letter(idx1+1)].width = value['width']
            else:
                titles.append(value)

        data_sheet.append(titles)

        for idx1, dataitem in enumerate(datalist):
            data = []
            # for idx2, key in enumerate(keys):

            if (dataitem['_height']):
                print('###############')
                data_sheet.row_dimensions[idx1 + 2].height = dataitem['_height']
                # data.append(idx1+1)
            for key in (keys):
                if (key == 'idx'):
                    data.append(idx1 + 1)
                print(key)

                # if(key in widthsmaps):

                #
                # sheet.column_dimensions[get_column_letter(idx)].width=50
                if (key in dataitem):
                    cellinfo = dataitem[key]
                    if (isinstance(cellinfo, dict)):
                        data.append(cellinfo['data'])
                    else:
                        data.append(cellinfo)

            # for idx2,key2 in enumerate(columnmaps.keys):
            data_sheet.append(data)
        # data = ['11', '22', '33', '44', '55', '66']
        # data_sheet.append(data)
        try:
            data_book.save(path)
        except Exception as ex:
            (file_name, extension) = os.path.splitext(path)
            print(file_name, extension)
            data_book.save(file_name+'_'+time.strftime('%Y%m%d%H%M%S')+extension)
    except Exception as ex:
        print(str(ex))

# def selectcolidx(workbook, workname, title):
#     worksheet = workbook.get_sheet_by_name(workname)
#     cols = worksheet.max_column
#     for item in range(cols):
#         if(title == worksheet.cell(1, item+1).value):
#             return item


class zyxlsx(unittest.TestCase):
    def test_updatexlsx(self):
        data_book = Workbook()
        updatexlsx(data_book, 'sheet1', ['序号:20', '标题', '类型',
                                         {'name': '图片', 'width': 20},
                   {'name': '差价比', 'formula': '=(B#/C#-1)', 'format': '0.00%'}], [{'标题': 111, '类型': 222}])

    # def test_writexls(self):
    #     writexls('zyxlsx.xls', [
    #         {'idx': '序号1', 'tp': '图片1', 'bt': '标题1', 'lx': '类型1'},
    #         {'idx': '序号2', 'tp': '图片2', 'bt': '标题2', 'lx': {'data': '类型1'}, '_height': 40}
    #     ], {
    #         '_idx': {'name': '序号', 'width': '20'},
    #         'idx': {'name': '序号', 'width': '20'},
    #         'tp': {'name': '图片', 'width': '20'},
    #         'bt': '标题',
    #         'lx': '类型'
    #     })


if __name__ == '__main__':
    unittest.main()
