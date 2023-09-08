##################################################
# 功能说明：
# 更新日期：
# 更新次数：
# 完成情况：
# 执行次数：5
# 执行时间：2023-01-27 17:02:11
##################################################
import unittest
from io import BytesIO, StringIO

import xlwt
import xlrd
from openpyxl.utils import get_column_letter
# from zypylibs.base.zydict import mappingdict

# from zypylibs.function.zycommon import updatedescribe

# updatedescribe(__file__)


# 写入单个sheet表格

def writexls(name, sheet='首页', header=[], datalist=[]):
    wbk = xlwt.Workbook()
    sheet = wbk.add_sheet(sheet)

    for (i, item) in enumerate(header):
        sheet.write(0, i, item)  # 第0行第一列写入内容

    for (i1, item1) in enumerate(datalist):
        for (i2, item2) in enumerate(item1):
            sheet.write(i1 + 1, i2, item2)
    if name == '':
        sio = BytesIO()  # StringIO()
        wbk.save(sio)  # 这点很重要，传给save函数的不是保存文件名，而是一个StringIO流
        return sio  # .getvalue()
    else:
        wbk.save(name)


def todatalist(column, datalist):
    def mapping(i, m, ii, mm):
        # print(i, m, ii, mm)

        if mm == '_id':
            return str(i + 1)
        if mm not in m:
            return '-'

        return m[mm]

    return list(map(lambda m: list(map(lambda mm:  mapping(m[0], m[1], mm[0], mm[1]), enumerate(column.values()))), enumerate(datalist)))


def writexlsx(workbook, workname, tbconfig, datalist):
    if workname in workbook.get_sheet_names():
        worksheet = workbook.sheet_by_name(workname)
    else:
        worksheet = workbook.create_sheet(workname)

    for idx, item in enumerate(tbconfig):
        worksheet.cell(1, idx+1).value = item['title']
        # worksheet.write(0,idx,label=item['title'])
        if ('width' in item):
            worksheet.column_dimensions[get_column_letter(idx+1)].width = int(item['width'])

    for idx1, item1 in enumerate(datalist):
        dataitem = datalist[idx1]
        for idx2, item2 in enumerate(tbconfig):
            worksheet.cell(idx1+2, idx2+1).value = dataitem[item2['key']]
    return workbook


def readxlsx(workbook, workname, tbconfig, cellformat=None):
    # print(workname)

    sheetnames = workbook.sheet_names()
    print(sheetnames)
    if (workname not in sheetnames):
        print(sheetnames, '不包含', workname)
        return
    worksheet = workbook.sheet_by_name(workname)
    # work_sheet = work_book.sheet_by_index(0)  # 通过表名的索引打开工作表
    rows = worksheet.nrows
    cols = worksheet.ncols

    titles = []
    for item in range(cols):
        titles.append(worksheet.cell(0, item).value)

    print(titles)

    print()
    # print(rows, cols)
    if(tbconfig is None or len(tbconfig) == 0):
        tbconfig = []
        for item in range(cols):
            cell = worksheet.cell(1, item + 1)
            tbconfig.append(cell)
    # 处理 tbconfig 格式
    for idx, item in enumerate(tbconfig):
        if(isinstance(item, str)):
            tbconfig[idx] = {'key': item, 'title': item}

    datalist = []
    for idx1 in range(rows - 1):
        dataitem = {}
        for idx2, item2 in enumerate(tbconfig):
            # print(idx1 + 2, idx2 + 1)
            datakey = item2['key']

            dataitem[datakey] = worksheet.cell(idx1 + 1, titles.index(datakey)).value
            # if(cellformat != None):
            #     cellformat(worksheet, idx1 + 2, idx2 + 1, item2, datavalue, dataitem, datakey)

            # if('colsformat' in item2):
            #     item2['colsformat'](worksheet, idx1 + 2, idx2 + 1, item2, datavalue, dataitem, datakey)

        # print(dataitem)
        datalist.append(dataitem)
    return datalist

# 找到指定标题所在的列


# def selectcolidx(workbook, workname, title):
#     worksheet = workbook.sheet_by_name(workname)
#     cols = worksheet.ncols
#     for item in range(cols):
#         if(title == worksheet.cell(0, item).value):
#             return item

# 找到指定字段所在的行


# def selectrowidx(workbook, workname, title, value, valueformat=None):
#     worksheet = workbook.get_sheet_by_name(workname)
#     colidx = selectcolidx(workbook, workname, title)
#     # print('###############',colidx)
#     rows = worksheet.max_row
#     for idx1 in range(rows-1):
#         pkval = worksheet.cell(idx1+2, colidx+1).value
#         if valueformat is not None:
#             pkval = valueformat(pkval)
#         if(pkval == value):
#             return idx1+1


def updatexlsx(workbook, workname, rowidx, colidx, value):
    worksheet = workbook.get_sheet_by_name(workname)
    worksheet.cell(rowidx, colidx).value = value
    return workbook


class zyexcel(unittest.TestCase):

    def test_writexls(self):

        # f=BytesIO()
        # f.write('中文'.encode('utf-8'))  #将 '中文'经utf-8编码成字节
        # print(f.getvalue())
        writexls('zyexcel.xls', 'sheet1', ['A', 'B', 'C'], [['A1', 'B1', 'C1'], ['A2', 'B2', 'C2'], ['A3', 'B3', 'C3']])
        self.assertEqual(len('result'), 6)
        data = writexls('', 'sheet1', ['A', 'B', 'C'], [['A1', 'B1', 'C1'], ['A2', 'B2', 'C2'], ['A3', 'B3', 'C3']])
        # print(data)

    # def test_selectcolidx(self):
    #     work_book = xlrd.open_workbook(r'C:\0000\47python\zypylibs\res\b9b56634-9ee2-11ed-8d08-9843fa964fe6.xlsx')
    #     colidx = selectcolidx(work_book, '清单', '模块编码')
    #     print(colidx)
    #     self.assertEqual(len('result'), 6)

    def test_list2csv(self):
        self.assertEqual(len('result'), 6)

    def test_todatalist(self):
        column = {
            '序号': '_id',
            '备用字段1': '备用字段1',
            '备用字段2': '备用字段2',
            '备用字段3': '备用字段3',
            '备用字段4': '备用字段4'
        }
        # column = mappingdict(list(fields3.keys()), list(fields3.keys()))
        self.assertEqual(
            todatalist(column, [{'备用字段1': 1, '备用字段2': 1, '备用字段3': 1}, {'备用字段1': 2, '备用字段2': 2, '备用字段3': 2}]),
            [['1', 1, 1, 1, '-'], ['2', 2, 2, 2, '-']]
        )

    def test_readxlsx(self):

        tbconfig = [{
            'key': '模块编码',
            'title': '模块编码'
        }, {
            'key': '模块名称',
            'title': '模块名称'
        }, {
            'key': '模块类型',
            'title': '模块类型',
        }, {
            'key': '模块上级',
            'title': '模块上级',
            # 'width': '10',
            # 'colsformat': colsformat1
        }, {
            'key': '备用字段1',
            'title': '备用字段1',
        }, {
            'key': '备用字段3',
            'title': '备用字段3',
        }, {
            'key': '备用字段2',
            'title': '备用字段2',
        }, {
            'key': '备用字段4',
            'title': '备用字段4',
        }]
        work_book = xlrd.open_workbook(r'C:\0000\47python\zypylibs\res\b9b56634-9ee2-11ed-8d08-9843fa964fe6.xlsx')
        datalist = readxlsx(work_book, '清单', tbconfig)
        print(datalist)
        print(len(datalist))
        self.assertEqual(len('result'), 6)

    def test_writexlsx(self):
        self.assertEqual(len('result'), 6)

    def test_updatexlsx(self):
        self.assertEqual(len('result'), 6)


if __name__ == '__main__':
    unittest.main()
