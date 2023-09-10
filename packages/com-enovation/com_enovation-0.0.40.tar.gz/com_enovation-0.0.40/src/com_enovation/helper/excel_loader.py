from logging import Logger, getLogger
from inspect import stack
from pandas import DataFrame
from openpyxl import load_workbook, Workbook

from com_enovation.helper.pandas_dataframe_typer import PandasDataframeTyper


class ExcelLoaderBean:
    _dict_tables: dict

    def __init__(
            self,
            dict_tables: dict,
    ):
        self._dict_tables = dict_tables

    @property
    def dict_tables(self) -> dict:
        return self._dict_tables


class ExcelLoader:
    _logger: Logger = getLogger(__name__)

    def load_tables(
            self,
            str_path: str,
            dict_tables_columns_types: dict = None
    ) -> ExcelLoaderBean:

        self._logger.debug(f"Function '{stack()[0].filename} - {stack()[0].function}' is called with parameters:"
                           f"\n\t- str_path: {str_path}")

        # read file
        obj_the_workbook: Workbook = load_workbook(filename=str_path, data_only=True)

        _dict_the_tables: dict = {}

        # For each sheet
        for i_sheet_name in obj_the_workbook.sheetnames:

            # We get the sheet
            obj_the_worksheet = obj_the_workbook[i_sheet_name]

            # For each table in this sheet
            for k_table_name, k_table_range in obj_the_worksheet.tables.items():

                # parse the data within the ref boundary
                obj_the_data = obj_the_worksheet[k_table_range]

                # extract the data
                # the inner list comprehension gets the values for each cell in the table
                lst_the_content: list = \
                    [
                        [
                            i_cell.value for i_cell in i_line
                        ]
                        for i_line in obj_the_data
                    ]

                lst_the_header: list = lst_the_content[0]

                # the contents ... excluding the header
                lst_the_lines: list = lst_the_content[1:]

                # create dataframe with the column names
                # and pair table name with dataframe
                df_the_dataframe = DataFrame(lst_the_lines, columns=lst_the_header)

                # We retype in case we have a dictionary
                if dict_tables_columns_types is not None:
                    if k_table_name in dict_tables_columns_types:
                        df_the_dataframe = PandasDataframeTyper.type(
                            df_to_type=df_the_dataframe,
                            dict_columns_to_type=dict_tables_columns_types[k_table_name]
                        )

                _dict_the_tables[k_table_name] = df_the_dataframe

        obj_the_return: ExcelLoaderBean = ExcelLoaderBean(dict_tables=_dict_the_tables)

        self._logger.debug(f"Function '{stack()[0].filename} - {stack()[0].function}' is returning.")

        return obj_the_return
