import os
from inspect import stack
from logging import getLogger, Logger
from pathlib import Path
from unittest import TestCase

import xlsxwriter
from openpyxl import load_workbook

from com_enovation.toolbox.excel_dashboard.excel_dashboarder import ExcelDashboarder
from com_enovation.toolbox.excel_dashboard.config_checker import ConfigChecker
from com_enovation.toolbox.excel_dashboard.widgets.cover_sheet.widget_cover_sheet import WidgetCoverSheet


class TestWidget(TestCase):

    _logger: Logger = getLogger(__name__)

    def test_01_config_checker(self):

        self._logger.debug(f"Function '{stack()[0].filename} - {stack()[0].function}' is called")

        _the_config_checker: ConfigChecker = ConfigChecker(
            lst_widgets_json_schemas=[WidgetCoverSheet.get_dict_config_schema()],
            str_base_json_schema_id=WidgetCoverSheet.get_widget_id()
        )

        _the_config_checker.validate(
            dict_to_validate={"narrative": "yo, this is fantastic {0} et pourquoi pas {1}"},
        )

        self._logger.debug(f"Function '{stack()[0].filename} - {stack()[0].function}' is returning")

    def test_02_unitary(self):

        self._logger.debug(f"Function '{stack()[0].filename} - {stack()[0].function}' is called")

        # We instantiate the widget
        _the_widget: WidgetCoverSheet = WidgetCoverSheet(
            str_address="B2",
            dict_parameters_map={"tokens": "jsg_tokens"},
            dict_config={"narrative": "yo, this is fantastic {0} et pourquoi pas {1}"},
            dict_raw_parameters={"jsg_tokens": ["ici c'est paris", "mais pourquoi donc?"]},
            str_sheet_name="abc",
            dict_default_format={}
        )

        # We print it into an excel spreadsheet
        _the_workbook = xlsxwriter.Workbook(Path(os.path.join(os.path.dirname(__file__), 'TC02.excel.xlsx')))
        _the_worksheet = _the_workbook.add_worksheet()

        # We print the widget
        _the_widget.write_to_excel(
            wb_workbook=_the_workbook,
            ws_worksheet=_the_worksheet
        )

        # We close the workbook
        _the_workbook.close()

        # We now read the excel spreadsheet, and assert the result is as expected

        _the_workbook = load_workbook(
            filename=Path(os.path.join(os.path.dirname(__file__), 'TC02.excel.xlsx')),
            data_only=True
        )
        _the_worksheet = _the_workbook.active
        _the_string = _the_worksheet["B2"].value

        self.assertEqual(
            first=_the_string,
            second=_the_widget.str_narrative.format(*_the_widget.lst_tokens),
            msg="The message stored in the excel spreadsheet is not as expected..."
        )

        self._logger.debug(f"Function '{stack()[0].filename} - {stack()[0].function}' is returning")

    def test_03_application(self):

        self._logger.debug(f"Function '{stack()[0].filename} - {stack()[0].function}' is called")

        _the_excel_dashboard: ExcelDashboarder = ExcelDashboarder()

        _the_excel_dashboard.excelize(
            p_output_file_path=Path(os.path.join(os.path.dirname(__file__), 'TC03.excel.xlsx')),
            dict_config={
                "sheets": {
                    "test_sheet_name": {
                        "widgets": {
                            "test_a_widget_label": {
                                "address": "C3",
                                "widget_id": "https://enovation.com/excel_dashboard/cover_sheet",
                                "config": {
                                    "narrative": "the first text is '{0}' and the second is '{1}'."
                                },
                            }
                        }
                    }
                }
            },
            **{"tokens": ["ici c'est paris", "mais pourquoi donc?"]}
        )

        # We now read the excel spreadsheet, and assert the result is as expected

        _the_workbook = load_workbook(
            filename=Path(os.path.join(os.path.dirname(__file__), 'TC03.excel.xlsx')),
            data_only=True
        )
        _the_worksheet = _the_workbook.active
        _the_string = _the_worksheet["C3"].value

        self.assertEqual(
            first=_the_string,
            second="the first text is '{0}' and the second is '{1}'.".format(
                *["ici c'est paris", "mais pourquoi donc?"]
            ),
            msg="The message stored in the excel spreadsheet is not as expected..."
        )

        self._logger.debug(f"Function '{stack()[0].filename} - {stack()[0].function}' is returned")
